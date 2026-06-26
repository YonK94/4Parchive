from __future__ import annotations

import glob
import json
import os
import re
from datetime import datetime
from typing import Optional, Union
from openpyxl import load_workbook
from ..config import Config
from .models import *


class PartService:
    
    def __init__(self):
        self.index_reader = IndexReader()

    def process_lot(
            self, lot_num: str, process: str,
            part_name: Optional[str] = None, include_ng: bool=True
            ) -> LotDataContainer:

        # 파일 경로 탐색
        if part_name:
            file_paths = self._find_files_by_list(lot_num, process, part_name)
            if not file_paths:
                raise FileNotFoundError("해당 LOT/공정의 데이터가 없습니다. (by list)")
            
        else:
            file_paths = self._find_files_by_keyword(lot_num, process)
            if not file_paths:
                single_file_path = self._find_single_file_in_lot(lot_num)
                if not single_file_path:
                    raise FileNotFoundError("해당 LOT/공정의 데이터가 없습니다. (by key)")
            part_name_keyword = self._extract_keyword_from_path(file_paths[0])
            part_name = self.index_reader.get_part_name(part_name_keyword)

        # 전처리
        lot_info, parser, inspector = self._preprocess_part(part_name, process, lot_num)

        # 데이터 처리 / 결과 모델 생성
        part_results = []
        count = 1
        for path in file_paths:
            # 경로 유효성 검사
            filename = os.path.basename(path)
            dirname = os.path.basename(os.path.dirname(path))
            if not (
                self._is_valid_filename(filename) and self._is_valid_dirname(dirname)
                ):
                # print(f"Exclude: 유효하지 않은 경로명 형식의 파일이 제외됨. ({path})")
                continue

            # S/N 및 작업자 추출 or 임의 생성
            serial_num: str = None
            worker: str = None
            if filename.count("_") == 1:
                serial_num = "--" + f"{lot_num[4:]}{count:03d}"
            else:
                file_base = os.path.splitext(filename)[0]
                serial_num = file_base.split("_")[3]
                worker = file_base.split("_")[4]

            # 값 추출 및 검사  # 이하 동일
            dims = parser.parse_dims(path)
            is_pass = inspector.inspect_dim(dims)

            # NG 필터링 옵션 처리
            if not include_ng and False in is_pass:
                # print(f"Exclude: NG가 포함된 파일이 제외됨. ({path})")
                continue
            
            # 결과 모델 생성
            result = PartResult(
                serial_num = serial_num,
                measured_time = self._generate_measured_time(filename),
                worked_cnc = "4-" + dirname[:2],
                worker = worker,
                dims = dims,
                is_pass = is_pass
            )
            part_results.append(result)
            count += 1

        return LotDataContainer(info = lot_info, parts = part_results)
    
    def process_each_part(
            self, paths: list[str], include_ng: bool=True
            ) -> LotDataContainer:
        for path in paths:
            # 파일명 유효성 검사 추가 소요
            # (KK20240701001_1차_2024년07월01일11시24분_01호_ABC123_김연용_D109-12992C(U1).xlsx)
            filename = os.path.basename(path)
            file_base = os.path.splitext(filename)[0]
            (lot_num, process, raw_datetime, raw_cnc,
             serial_num, worker, part_name) = file_base.split('_')
            
            lot_info, parser, inspector = self._preprocess_part(part_name, process, lot_num)
        # 5.

    def _preprocess_part(
            self, part_name: str, process: str, lot_num: str
            ) -> tuple[LotInfo, ExcelParser, DimInspector]:
        # 규격 및 기준 정보 로드
        prod_info = self.index_reader.get_product_info(part_name, process)
        basic_data = prod_info["basic_data"]

        # LotInfo 공통 모델 생성
        lot_info = LotInfo(
            lot_num = lot_num,
            process = process,
            part_name = part_name,
            dim_basic_name = basic_data.dim_name,
            dim_basic_float = basic_data.dim_float,
            tol_upper = basic_data.upper,
            tol_lower = basic_data.lower
            # dim_upper = [b + t for b, t in zip(basic_data.dim_float, basic_data.upper)],
            # dim_lower = [b + l for b, l in zip(basic_data.dim_float, basic_data.lower)]
        )

        # 파서/검사기 초기화
        parser = ExcelParser(basic_data.cell)
        inspector = DimInspector(
            lot_info.dim_basic_float,
            lot_info.tol_upper,
            lot_info.tol_lower
            )
        
        return lot_info, parser, inspector

    def _find_files_by_list(
            self, lot_num: str, process: str, part_name :str
            ) -> List[str]:
        dir_name_list = self.index_reader.get_data_path(part_name, process)
        paths = []
        for dir in dir_name_list:
            search_pattern = fr"{Config.DATA_ROOT}\*{dir}\{lot_num}*"
            paths.extend(glob.glob(search_pattern))
        return paths

    def _find_files_by_keyword(self, lot_num: str, process: str) -> List[str]:
        if process == "가공":
            search_pattern = fr"{Config.DATA_ROOT}\*호*\{lot_num}*"
        else:
            search_pattern = fr"{Config.DATA_ROOT}\*호*{process}\{lot_num}*"
        return glob.glob(search_pattern)

    def _find_single_file_in_lot(self, lot_num: str) -> List[str]:
        file_gen = glob.iglob(fr"{Config.DATA_ROOT}\*호*\{lot_num}*", recursive=True)
        first_file = next(file_gen, None)
        return first_file

    def _extract_keyword_from_path(self, path: str) -> str:
        dirname = os.path.basename(os.path.dirname(path))
        dirname_splited = dirname.split()
        if dirname_splited[-1].endswith(("차", "삭", "통합")):
            return " ".join(dirname_splited[1:-1])
        else:
            return " ".join(dirname_splited[1:])
    
    def _is_valid_dirname(self, dirname: str) -> bool:
        return bool(re.match(Config.PATTERN_DIRNAME, dirname))

    def _is_valid_filename(self, filename: str) -> bool:
        return bool(
            re.match(
                Config.PATTERN_FILENAME_1, filename
                ) or re.match(Config.PATTERN_FILENAME_2, filename)
            )

    def _generate_measured_time(self, raw_str: str) -> str:
        format_string = "%Y년%m월%d일%H시%M분%S초"
        if "_" in raw_str:
            file_base = os.path.splitext(raw_str)[0]   # 확장자 제거.    
            raw_str = file_base.split("_")[1]    # Lot 번호 제거.
        dt_object = datetime.strptime(raw_str, format_string)
        return dt_object.strftime("%Y-%m-%d %H:%M:%S")
    

class IndexReader:
    def __init__(self):
        with open(Config.FILE_MAP_PART_NAME, 'r', encoding='utf-8') as f:
            self._map_part_name = json.load(f)
        with open(Config.FILE_INDEX_PART_DATA, 'r', encoding='utf-8') as f:
            self._index_part_data = json.load(f)
        with open(Config.FILE_MAP_DATA_PATH, 'r', encoding='utf-8') as f:
            self._map_data_path = json.load(f)
            
    def get_part_name(self, keyword: str) -> str:
        part_name = self._map_part_name.get(keyword)
        if part_name == None:
            raise KeyError(f"map_part_name.json에 '{keyword}'이(가) 없습니다.")
        return self._map_part_name.get(keyword)
    
    def get_data_path(self, part_name: str, process: str) -> list[str]:
        process_list = self._map_data_path.get(part_name)
        if not process_list:
            raise KeyError(f"map_data_path.json에 등록되지 않은 규격입니다.")
        dir_name_list = process_list.get(process)
        if not dir_name_list:
            raise KeyError(f"map_data_path.json에 등록되지 않은 공정입니다.")
        return dir_name_list

    def get_product_info(self, part_name: str, process: str) -> ProductInfo:
        data = self._index_part_data.get(part_name)
        if not data:
            raise KeyError(f"index_part_data.json에 '{part_name}'이(가) 없습니다.")
        elif data.get(process) is None:
            raise KeyError(f"index_part_data.json에 '{process}' 공정이 없습니다.")
        return {
            # "product_name": data["product"],
            "basic_data": BasicDataSpecs(**data[process])
        }


class DimInspector:
    def __init__(
            self,
            dim_basic: list[float],
            tol_upper: list[float],
            tol_lower: list[float]
            ):
        self.dim_basic = dim_basic
        self.dim_upper = [b + t for b, t in zip(dim_basic, tol_upper)]
        self.dim_lower = [b + t for b, t in zip(dim_basic, tol_lower)]

    def inspect_dim(self, dims: list[float]) -> list[bool]:
        results = []
        for i, dim in enumerate(dims):
            if dim == None:
                results.append(None)
            else:
                if self.dim_lower[i] <= dim <= self.dim_upper[i]:
                    results.append(True)
                else:
                    results.append(False)
        return results
    

class ExcelParser:
    def __init__(self, cell_addresses: List[str]):
        self.cell_addresses = cell_addresses

    def parse_dims(self, file_path: str) -> List[Optional[float]]:
        ext = os.path.splitext(file_path)[1].lower()
        dims = []
        
        try:
            if ext == ".xlsx":
                dims = self._parse_xlsx(file_path)
            else:
                print(f"Error: 지원하지 않는 파일 형식입니다. ({file_path})")
                return [None] * len(self.cell_addresses)
            
        except Exception as e:
            print(f"Error: 파일 파싱 실패 ({file_path}) - {e}")
            return [None] * len(self.cell_addresses)
            
        return dims

    def _parse_xlsx(self, path: str) -> List[Optional[float]]:
        wb = load_workbook(path, read_only = True, data_only = True)
        ws = wb.active
        
        dims = []
        for addr in self.cell_addresses:
            try:
                cell = ws[addr]
                val = cell.value
                dims.append(self._safe_float(val))
            except:
                dims.append(None)

        wb.close()
        return dims

    def _safe_float(
            self, value: Union[str, int, float, None]
            ) -> Optional[float]:
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
            
        if isinstance(value, str):
            value = value.strip() # 공백 제거
            if not value:
                return None
            try:
                return float(value)
            except ValueError:
                return None # 문자가 있는 경우
                
        return None